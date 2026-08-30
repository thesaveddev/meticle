#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()

# Colors
NAVY = "0F4C81"
WHITE = "FFFFFF"
LIGHT_BG = "F8FAFC"
CRITICAL_RED = "DC2626"
HIGH_ORANGE = "D97706"
MEDIUM_BLUE = "2563EB"
LOW_GREEN = "16A34A"
GREEN_BG = "DCFCE7"
BLUE_BG = "DBEAFE"
RED_BG = "FEE2E2"
GREY_BG = "F1F5F9"

header_font = Font(bold=True, color=WHITE, size=11)
header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
header_align = Alignment(vertical="center", horizontal="center", wrap_text=True)
thin_border = Border(bottom=Side(style="thin", color="E2E8F0"))

# ═══════════════════════════════════════════════
# Sheet 1: Go-Live Readiness Checklist
# ═══════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Go-Live Readiness"
ws1.sheet_properties.tabColor = NAVY

headers = ["ID", "Priority", "Category", "Item", "Description", "Owner", "Effort", "Status", "% Complete", "Target Date", "Notes / Blockers"]
widths = [8, 14, 18, 42, 55, 16, 14, 14, 12, 14, 40]

for i, (h, w) in enumerate(zip(headers, widths), 1):
    cell = ws1.cell(row=1, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.freeze_panes = "A2"

items = [
    # CRITICAL
    ("CRIT-01", "🔴 CRITICAL", "Security", "Penetration test", "Hire CREST/CHECK-certified pen tester to test auth bypass, IDOR, SQL injection, XSS, API rate limiting, health data leakage", "", "1-2 weeks", "Not Started", 0, "", "Budget ~£3-5k"),
    ("CRIT-02", "🔴 CRITICAL", "Legal / GDPR", "Data Protection Impact Assessment (DPIA)", "Required under UK GDPR Article 35 for processing special category health data at scale. Must cover lawful basis, data minimisation, retention, international transfers, AI automated decision-making", "", "1 week", "Not Started", 0, "", ""),
    ("CRIT-03", "🔴 CRITICAL", "QA", "Run full QA test pack", "Execute all 381 test cases in docs/MeticleCare_QA_UAT_Test_Pack.xlsx. Fill in Pass/Fail and Comments. Return for bug fixes", "", "3-5 days", "Not Started", 0, "", ""),
    ("CRIT-04", "🔴 CRITICAL", "Security", "SSL certificates verified", "Verify production has valid SSL certs with auto-renewal configured. Test HTTPS on all endpoints", "", "1 hour", "Not Started", 0, "", ""),
    ("CRIT-05", "🔴 CRITICAL", "Data", "Backup & recovery testing", "Perform test database restore from backup. Verify data integrity. Document RTO/RPO. Test point-in-time recovery", "", "1 day", "Not Started", 0, "", ""),
    ("CRIT-06", "🔴 CRITICAL", "Infrastructure", "Database connection pooling verified", "Verify PgBouncer or equivalent is configured. Test connection pool exhaustion under load. Set up connection pool monitoring", "", "Half day", "Not Started", 0, "", ""),
    ("CRIT-07", "🔴 CRITICAL", "Legal / GDPR", "DPIA reviewed by DPO or legal counsel", "The DPIA document must be reviewed and signed off by a qualified Data Protection Officer or solicitor before launch", "", "1-2 days", "Not Started", 0, "", ""),
    # HIGH
    ("HIGH-01", "🟠 HIGH", "Legal / GDPR", "ICO registration", "Register as data controller with ICO (ico.org.uk). Required if processing personal data. Currently £40-60/year for small orgs", "", "1 hour", "Not Started", 0, "", ""),
    ("HIGH-02", "🟠 HIGH", "Monitoring", "Uptime monitoring & alerting", "Set up uptime monitoring (UptimeRobot, Betterstack, or similar). Configure alerts for API downtime, high error rates, database issues, disk space", "", "2-3 hours", "Not Started", 0, "", ""),
    ("HIGH-03", "🟠 HIGH", "Legal", "Solicitor review — Terms of Use", "Have a solicitor review Terms of Use for enforceability, limitation of liability, indemnification clauses specific to health/social care software", "", "1-2 weeks", "Not Started", 0, "", ""),
    ("HIGH-04", "🟠 HIGH", "Legal", "Solicitor review — Privacy Policy", "Review Privacy Policy for UK GDPR compliance. Ensure lawful basis for processing health data is clearly documented. Verify data subject rights are covered", "", "1-2 weeks", "Not Started", 0, "", ""),
    ("HIGH-05", "🟠 HIGH", "Legal", "Create standalone DPA document", "The Terms of Use references a Data Processing Agreement. Create a standalone DPA for care homes signing up as data controllers", "", "3-5 days", "Not Started", 0, "", ""),
    ("HIGH-06", "🟠 HIGH", "Legal", "Cookie policy accuracy review", "Verify Cookie Policy accurately reflects all cookies used. Update if any analytics or tracking is added post-launch", "", "2 hours", "Not Started", 0, "", ""),
    ("HIGH-07", "🟠 HIGH", "Insurance", "Professional indemnity insurance", "Obtain PI insurance covering software defects that could cause harm in a care setting. Also consider cyber liability insurance for data breach", "", "1 week", "Not Started", 0, "", ""),
    ("HIGH-08", "🟠 HIGH", "Infrastructure", "Redis configured for Socket.IO", "Verify Redis adapter is configured for horizontal scaling of real-time features (chat, notifications, mission control). Test with multiple API instances", "", "Half day", "Not Started", 0, "", ""),
    ("HIGH-09", "🟠 HIGH", "Security", "API key rotation policy", "Implement or document a policy for rotating API keys (OpenAI, Anthropic, Stripe). Set calendar reminders for quarterly rotation", "", "2 hours", "Not Started", 0, "", ""),
    ("HIGH-10", "🟠 HIGH", "Security", "Rate limiting verified in production", "Verify rate limits are applied to all public-facing endpoints. Test with concurrent requests to confirm 429 responses", "", "2 hours", "Not Started", 0, "", ""),
    # MEDIUM
    ("MED-01", "🟡 MEDIUM", "Compliance", "CQC registration guidance", "Clarify whether MeticleCare needs CQC registration (only if providing care directly) or if it is purely software for care providers", "", "Research", "Not Started", 0, "", ""),
    ("MED-02", "🟡 MEDIUM", "Operations", "Incident response plan", "Document what happens when: data breach occurs, system goes down, safeguarding concern arises. Who to call, what to do, comms templates", "", "2-3 days", "Not Started", 0, "", ""),
    ("MED-03", "🟡 MEDIUM", "Operations", "Staff onboarding documentation", "Create internal docs for any team members who will support the platform: how to access logs, how to reset users, how to check billing", "", "1-2 days", "Not Started", 0, "", ""),
    ("MED-04", "🟡 MEDIUM", "Infrastructure", "Production environment audit", "Review docker-compose.prod.yml, verify all services have health checks, proper resource limits, log rotation, and restart policies", "", "Half day", "Not Started", 0, "", ""),
    ("MED-05", "🟡 MEDIUM", "Data", "Data retention policy implementation", "Ensure the app enforces data retention periods. Health data retention per NHS records management code of practice. Implement auto-deletion or archival", "", "2-3 days", "Not Started", 0, "", ""),
    ("MED-06", "🟡 MEDIUM", "AI", "AI content labelling audit", "Verify all AI-generated outputs (daily notes, meal plans, incident triage, compliance analysis) are clearly labelled as AI-generated in the UI and database", "", "1 day", "Not Started", 0, "", ""),
    ("MED-07", "🟡 MEDIUM", "Operations", "Disaster recovery runbook", "Document step-by-step recovery procedures: database restore, service restart, DNS failover, data integrity checks. Test the runbook", "", "1-2 days", "Not Started", 0, "", ""),
    ("MED-08", "🟡 MEDIUM", "Accessibility", "WCAG 2.1 AA compliance audit", "Run axe-core or similar against key pages. Fix critical violations (missing alt text, colour contrast, keyboard navigation, screen reader support)", "", "2-3 days", "Not Started", 0, "", ""),
    # LOW
    ("LOW-01", "🟢 LOW", "Performance", "Load testing", "Run k6 or Artillery against the API. Simulate 50, 100, 200 concurrent users. Measure response times, error rates, and resource usage", "", "1-2 days", "Not Started", 0, "", ""),
    ("LOW-02", "🟢 LOW", "Infrastructure", "Domain & DNS setup", "Finalise production domain. Configure DNS records. Set up email sending (SPF/DKIM/DMARC for transactional emails)", "", "2-3 hours", "Not Started", 0, "", ""),
    ("LOW-03", "🟢 LOW", "Operations", "Customer support process", "Define how support tickets are handled: email, chat, phone. Who responds. SLA targets. Escalation path for urgent issues", "", "Half day", "Not Started", 0, "", ""),
    ("LOW-04", "🟢 LOW", "Compliance", "CQC evidence pack auto-generation test", "Verify evidence packs generate correctly with real data. Test all CQC domains (Safe, Effective, Caring, Responsive, Well-led)", "", "Half day", "Not Started", 0, "", ""),
    ("LOW-05", "🟢 LOW", "Marketing", "Landing page A/B test plan", "Plan initial A/B tests for the landing page: headline variations, CTA placement, pricing display, trust strip positioning", "", "Half day", "Not Started", 0, "", ""),
    ("LOW-06", "🟢 LOW", "Documentation", "API documentation", "Generate or write API docs for any public-facing or integration APIs. Consider OpenAPI/Swagger spec generation", "", "2-3 days", "Not Started", 0, "", ""),
    ("LOW-07", "🟢 LOW", "Marketing", "Social proof & testimonials", "Collect early beta user feedback. Create case studies or testimonials. Add to landing page once available", "", "Ongoing", "Not Started", 0, "", ""),
]

prio_fonts = {
    "🔴 CRITICAL": Font(bold=True, color=CRITICAL_RED),
    "🟠 HIGH": Font(bold=True, color=HIGH_ORANGE),
    "🟡 MEDIUM": Font(bold=True, color=MEDIUM_BLUE),
    "🟢 LOW": Font(bold=True, color=LOW_GREEN),
}
alt_fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")

for idx, item in enumerate(items):
    row = idx + 2
    for col, val in enumerate(item, 1):
        cell = ws1.cell(row=row, column=col, value=val)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = thin_border
        if idx % 2 == 0:
            cell.fill = alt_fill
    # Priority font
    ws1.cell(row=row, column=2).font = prio_fonts.get(item[1], Font())

# Status dropdown
status_dv = DataValidation(type="list", formula1='"Not Started,In Progress,Blocked,Complete,N/A"', allow_blank=True)
status_dv.error = "Please select a valid status"
status_dv.errorTitle = "Invalid Status"
ws1.add_data_validation(status_dv)
status_dv.add(f"H2:H{len(items)+1}")

# Effort dropdown
effort_dv = DataValidation(type="list", formula1='"1 hour,Half day,1 day,2-3 days,3-5 days,1 week,1-2 weeks,2-3 weeks,Ongoing,Research"', allow_blank=True)
ws1.add_data_validation(effort_dv)
effort_dv.add(f"G2:G{len(items)+1}")

# Conditional formatting for Status
ws1.conditional_formatting.add(f"H2:H{len(items)+1}", CellIsRule(operator="equal", formula=['"Complete"'], fill=PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid"), font=Font(bold=True, color=LOW_GREEN)))
ws1.conditional_formatting.add(f"H2:H{len(items)+1}", CellIsRule(operator="equal", formula=['"In Progress"'], fill=PatternFill(start_color=BLUE_BG, end_color=BLUE_BG, fill_type="solid"), font=Font(bold=True, color=MEDIUM_BLUE)))
ws1.conditional_formatting.add(f"H2:H{len(items)+1}", CellIsRule(operator="equal", formula=['"Blocked"'], fill=PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid"), font=Font(bold=True, color=CRITICAL_RED)))
ws1.conditional_formatting.add(f"H2:H{len(items)+1}", CellIsRule(operator="equal", formula=['"Not Started"'], fill=PatternFill(start_color=GREY_BG, end_color=GREY_BG, fill_type="solid"), font=Font(color="94A3B8")))

# ═══════════════════════════════════════════════
# Sheet 2: Summary
# ═══════════════════════════════════════════════
ws2 = wb.create_sheet("Summary")
ws2.sheet_properties.tabColor = "2563EB"

sum_headers = ["Priority", "Total", "Not Started", "In Progress", "Blocked", "Complete", "% Complete"]
for i, h in enumerate(sum_headers, 1):
    cell = ws2.cell(row=1, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    ws2.column_dimensions[get_column_letter(i)].width = 16

priorities = ["🔴 CRITICAL", "🟠 HIGH", "🟡 MEDIUM", "🟢 LOW"]
for p_idx, p in enumerate(priorities):
    row = p_idx + 2
    count = sum(1 for i in items if i[1] == p)
    ws2.cell(row=row, column=1, value=p).font = prio_fonts[p]
    ws2.cell(row=row, column=2, value=count)
    ws2.cell(row=row, column=3, value=count)  # Not Started
    for c in range(4, 7):
        ws2.cell(row=row, column=c, value=0)
    ws2.cell(row=row, column=7, value="0%")

# Total row
total_row = len(priorities) + 2
bold_font = Font(bold=True, size=11)
top_border = Border(top=Side(style="medium", color=NAVY))
ws2.cell(row=total_row, column=1, value="TOTAL").font = bold_font
ws2.cell(row=total_row, column=2, value=len(items)).font = bold_font
ws2.cell(row=total_row, column=3, value=len(items)).font = bold_font
for c in range(1, 8):
    ws2.cell(row=total_row, column=c).border = top_border
    ws2.cell(row=total_row, column=c).font = bold_font

# ═══════════════════════════════════════════════
# Sheet 3: By Category
# ═══════════════════════════════════════════════
ws3 = wb.create_sheet("By Category")
ws3.sheet_properties.tabColor = "16A34A"

cat_headers = ["Category", "Total Items", "Critical", "High", "Medium", "Low"]
for i, h in enumerate(cat_headers, 1):
    cell = ws3.cell(row=1, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    ws3.column_dimensions[get_column_letter(i)].width = 18

cats = list(dict.fromkeys(i[2] for i in items))  # preserve order, dedupe
for c_idx, cat in enumerate(cats):
    row = c_idx + 2
    cat_items = [i for i in items if i[2] == cat]
    ws3.cell(row=row, column=1, value=cat).font = Font(bold=True)
    ws3.cell(row=row, column=2, value=len(cat_items))
    ws3.cell(row=row, column=3, value=sum(1 for i in cat_items if "CRITICAL" in i[1]))
    ws3.cell(row=row, column=4, value=sum(1 for i in cat_items if "HIGH" in i[1]))
    ws3.cell(row=row, column=5, value=sum(1 for i in cat_items if "MEDIUM" in i[1]))
    ws3.cell(row=row, column=6, value=sum(1 for i in cat_items if "LOW" in i[1]))
    for c in range(1, 7):
        ws3.cell(row=row, column=c).border = thin_border

# ═══════════════════════════════════════════════
# Sheet 4: Timeline
# ═══════════════════════════════════════════════
ws4 = wb.create_sheet("Timeline")
ws4.sheet_properties.tabColor = "D97706"

tl_headers = ["Week", "Focus Area", "Key Items", "Status", "Notes"]
tl_widths = [12, 35, 65, 14, 40]
for i, (h, w) in enumerate(zip(tl_headers, tl_widths), 1):
    cell = ws4.cell(row=1, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    ws4.column_dimensions[get_column_letter(i)].width = w

timeline = [
    ("Week 1", "Security & Legal", "Engage pen tester, begin DPIA draft, ICO registration, SSL verification", "Not Started", ""),
    ("Week 2", "QA & Testing", "Run full QA test pack (381 cases), document all failures, begin load testing", "Not Started", ""),
    ("Week 3", "Legal & Compliance", "DPIA review by DPO, solicitor review of Terms + Privacy Policy, create DPA document", "Not Started", ""),
    ("Week 4", "Infrastructure & Monitoring", "Uptime monitoring setup, backup recovery testing, Redis scaling test, prod audit", "Not Started", ""),
    ("Week 5", "Bug Fixes & Polish", "Fix all QA failures, address pen test findings, AI labelling audit, WCAG fixes", "Not Started", ""),
    ("Week 6", "Launch Prep", "Final smoke test, incident response runbook, monitoring live, disaster recovery test", "Not Started", ""),
]

for t_idx, tl in enumerate(timeline):
    row = t_idx + 2
    for col, val in enumerate(tl, 1):
        cell = ws4.cell(row=row, column=col, value=val)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = thin_border

# Save
out_dir = os.path.join(os.getcwd(), "docs")
os.makedirs(out_dir, exist_ok=True)
out_path = os.path.join(out_dir, "MeticleCare_GoLive_Readiness.xlsx")
wb.save(out_path)
print(f"Written: {out_path}")
print(f"   {len(items)} items across {len(priorities)} priority levels")
print(f"   4 sheets: Go-Live Readiness, Summary, By Category, Timeline")
